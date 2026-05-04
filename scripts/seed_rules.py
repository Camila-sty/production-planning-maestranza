#!/usr/bin/env python3
"""
Seed script: reads /data/Reglas_Planificación.xlsx and inserts data into SQLite DB.
Usage: python3 scripts/seed_rules.py [db_path]
Default db_path: etp-app/dev.db
"""
import sys
import os
import sqlite3
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# Resolve paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH = os.path.join(PROJECT_ROOT, "data", "Reglas_Planificación.xlsx")
DB_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.join(PROJECT_ROOT, "etp-app", "dev.db")

def cuid_like(prefix: str, unique: str) -> str:
    import hashlib
    h = hashlib.sha256(f"{prefix}:{unique}".encode()).hexdigest()[:20]
    return f"c{h}"

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at: {EXCEL_PATH}")
        sys.exit(1)

    if not os.path.exists(DB_PATH):
        print(f"ERROR: Database not found at: {DB_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # --- Sheet 1: Lead times by codigo_plazo ---
    sheet_lt = wb["CÓDIGO PLAZO Y DEMORAS EN DÍAS"]
    rows_lt = list(sheet_lt.iter_rows(values_only=True))

    header = rows_lt[0]  # ('CÓDIGO PLAZO', 'DESCRIPCIÓN EQUIPO', 'INGENIERÍA', ...)
    process_cols = list(header[2:])  # column names starting from index 2

    lead_times = []
    for row in rows_lt[1:]:
        if row[0] is None:
            continue
        codigo_plazo = str(int(row[0]))
        descripcion = str(row[1]) if row[1] else None
        for col_idx, proc_name in enumerate(process_cols):
            if proc_name is None:
                continue
            duracion = row[2 + col_idx]
            if duracion is None:
                duracion = 0
            lead_times.append({
                "codigo_plazo": codigo_plazo,
                "descripcion_equipo": descripcion,
                "proceso": str(proc_name).strip(),
                "duracion_dias": int(duracion),
            })

    # --- Sheet 2: Process capacities ---
    sheet_pc = wb["CAPACIDAD POR PROCESO"]
    rows_pc = list(sheet_pc.iter_rows(values_only=True))

    process_capacities = []
    for row in rows_pc[1:]:
        if row[0] is None:
            continue
        proc_name = str(row[0]).strip()
        orden = int(row[1]) if row[1] is not None else 0
        capacidad = int(row[2]) if row[2] is not None else 0
        process_capacities.append({
            "proceso": proc_name,
            "orden": orden,
            "capacidad_por_dia": capacidad,
        })

    # --- Insert into DB ---
    conn = sqlite3.connect(DB_PATH)
    now = datetime.utcnow().isoformat()

    # Upsert lead_time_by_code
    inserted_lt = 0
    updated_lt = 0
    for lt in lead_times:
        row_id = cuid_like("lt", f"{lt['codigo_plazo']}:{lt['proceso']}")
        existing = conn.execute(
            "SELECT id FROM lead_time_by_code WHERE codigo_plazo=? AND proceso=?",
            (lt["codigo_plazo"], lt["proceso"])
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE lead_time_by_code SET descripcion_equipo=?, duracion_dias=?, updated_at=? WHERE id=?",
                (lt["descripcion_equipo"], lt["duracion_dias"], now, existing[0])
            )
            updated_lt += 1
        else:
            conn.execute(
                "INSERT INTO lead_time_by_code (id, codigo_plazo, descripcion_equipo, proceso, duracion_dias, created_at, updated_at) VALUES (?,?,?,?,?,?,?)",
                (row_id, lt["codigo_plazo"], lt["descripcion_equipo"], lt["proceso"], lt["duracion_dias"], now, now)
            )
            inserted_lt += 1

    # Upsert process_capacity
    inserted_pc = 0
    updated_pc = 0
    for pc in process_capacities:
        row_id = cuid_like("pc", pc["proceso"])
        existing = conn.execute(
            "SELECT id FROM process_capacity WHERE proceso=?",
            (pc["proceso"],)
        ).fetchone()
        if existing:
            conn.execute(
                "UPDATE process_capacity SET orden=?, capacidad_por_dia=?, updated_at=? WHERE id=?",
                (pc["orden"], pc["capacidad_por_dia"], now, existing[0])
            )
            updated_pc += 1
        else:
            conn.execute(
                "INSERT INTO process_capacity (id, proceso, orden, capacidad_por_dia, created_at, updated_at) VALUES (?,?,?,?,?,?)",
                (row_id, pc["proceso"], pc["orden"], pc["capacidad_por_dia"], now, now)
            )
            inserted_pc += 1

    conn.commit()
    conn.close()

    print(f"LeadTimeByCode:  {inserted_lt} inserted, {updated_lt} updated")
    print(f"ProcessCapacity: {inserted_pc} inserted, {updated_pc} updated")
    print("Seed complete.")

if __name__ == "__main__":
    main()
